import tkinter as tk
from tkinter import*
from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
import re
from tkcalendar import DateEntry
from tkcalendar import*
from tkinter import font
from tkinter.ttk import Combobox
from openpyxl import load_workbook
import pathlib
import openpyxl



framebg='white'
framefg='#326273'
background='#f0f0ed'

root=tk.Tk()
root.title("Student_Data")
root.configure(bg="#326273")
root.geometry("1200x700+210+100")
root.config(bg=background)
root.rowconfigure(0, weight=1)
root.columnconfigure(0, weight=1)
root.state('zoomed')


excel_path=r"Student_Info.xlsx"

file=pathlib.Path('Student_Info.xlsx')
if file.exists():
    pass
else:
    file=load_workbook()
    sheet=file.active

    sheet['A1']="Register No:"
    sheet['B1']="Name:"
    sheet['C1']="Academic Year"
    sheet['D1']="Class:"
    sheet['E1']="Gender:"
    sheet['F1']="DOB"
    sheet['G1']="Mobile No:"
    sheet['H1']="Mail ID:"
    sheet['I1']="Blood  Group :"
    sheet['J1']="Aadhar No :"
    sheet['K1']="Father Name:"
    sheet['L1']="Father's Occupation :"
    sheet['M1']="Mother Name:"
    sheet['N1']="Mother's Occupation :"
    sheet['O1']="Parent's Mobile No :"
    sheet['P1']="Mother Tongue:"
    sheet['Q1']="Religion:"
    sheet['R1']="Community:"
    sheet['S1']="Sub Cast:"
    sheet['T1']="Income:"
    sheet['U1']="Communication Add"
    sheet['V1']="Permanent Add"
    sheet['W1']="School name"
    sheet['X1']="School Type"
    sheet['Y1']="Medium"
    sheet['Z1']="Board"
    sheet['AA1']="12th Mark"
    sheet['AB1']="Percentage"
    sheet['AC1']="NO of Attempt"
    sheet['AD1']="Passing month"
    sheet['AE1']="Passing year"
    sheet['AF1']="EMIS NO"
    sheet['AG1']="Bank Name"
    sheet['AH1']="Bank Branch"
    sheet['AI1']="Account NO"
    sheet['AJ1']="IFSC Code"
    sheet['AK1']="MICR Code"
    sheet['AL1']="UMIS No"
    sheet['AM1']="NM ID"
    sheet['AN1']="PP ID"
    sheet['AO1']="TP ID"

    file.save('Student_info.xlsx')


icon_image=PhotoImage(file="bcaicon.png")
root.iconphoto(True,icon_image)

page1 =Frame(root)
page2 =Frame(root)
page3 =Frame(root)

for frame in (page1, page2, page3):
  frame.grid(row=0, column=0, sticky='nsew')

def show_frame(frame):
  frame.tkraise()

show_frame(page1)

#===========================PAGE 1 ======================
"""
img1=PhotoImage(file="images/BCA logo.png")
lbl2=Label(bg="#FFFF00",image=img1)
lbl2.place(x=570,y=180)"""

label =Label(page1,width=10,height=9,bg="#000066",anchor='e').pack(side=TOP,fill=X)
label =Label(page1,text="WELCOME",width=10,height=0,bg="#000066",fg='#FFFF00',font='arial 50 bold').place(x=480,y=50)
label =Label(page1,width=10,height=50,bg='#FFFF00').pack(side=TOP,fill=X)
label =Label(page1,text="BCA",width=10,height=0,bg="#FFFF00",fg='#003399',font='arial 60 bold').place(x=435,y=200)
label =Label(page1,text="Bachelor of Computer Applications",width=0,height=0,bg='#FFFF00',fg='#000066',font='arial 12 bold').place(x=540,y=300)
label =Label(page1,text="Student Registration System",width=0,height=0,bg='#FFFF00',fg='#000066',font='arial 30 bold').place(x=420,y=330)
label =Label(page1,text="Main Menu",width=0,height=0,bg='#FFFF00',fg='#003399',font='arial 20 bold',anchor='e').place(x=610,y=430)

"""====================================="""
  
page1_button= Button(page1,text="New Form >",width=19,height=2,font="Helvetica 12 bold",bg="#00FF00",bd=2, command=lambda: show_frame(page2))
page1_button.place(x=465,y=480)

def open_excel_file():

    file_path = filedialog.askopenfilename(title="Select Excel File",filetypes=[("Excel files","*.xls *.xlsx *.csv")])
    if file_path:
        os.startfile(file_path)

page1_button= Button(page1,text="Existing File >",width=19,height=2,font="Helvetica 12 bold",bg="#00CCFF",bd=2, command=open_excel_file)
page1_button.place(x=690,y=480)

def Exit():
    root.destroy()

Button(page1,text="EXIT",width=19,height=2,font="arial 12 bold",bg="Deep Pink",bd=1,command=Exit).place(x=580,y=650)


#================ page 2 =================================

page2_label =Label(page2,text="Email : kamalcrush15@gmail.com ",width=10,height=3,bg="#778899",anchor='e').pack(side=TOP,fill=X)
page2_label =Label(page2,text="STUDENT REGISTRATION ",width=10,height=2,bg='#326273',fg='#fff',font='arial 20 bold').pack(side=TOP,fill=X)

def submit():

    
  
  register=Register.get()
  name=Name.get()
  accyear=Accademicyear.get()
  Class=Class_Combobox.get()
  gender=gender_Combobox.get()
  dob=DOB.get()
  Mobile=Mobile_No.get()
  Mail=Mail_Id.get()
  Blood=B_group.get()
  Aadhaar=Aadhar.get()
  Father=F_Name.get()
  FOccupation=F_Occupation.get()
  Mother=M_Name.get()
  MOccupation=M_Occupation.get()
  ParentNo=Parent_No.get()
  Mtongue=M_tongue.get()
  religion=Religion.get()
  community=Community.get()
  SubCast=Sub_Cast.get()
  income=Income.get()
  C_Address=C_AddressEntry.get(1.0,END)
  P_Address=P_AddressEntry.get(1.0,END)
  School=S_Name.get()
  S_Type=S_Type_Combobox.get()
  medium=Medium.get()
  Board=Board_Combobox.get()
  mark=Mark.get()
  percentage=Percentage.get()
  attempt=Attempt.get()
  Month=P_Month.get()
  year=P_year.get()
  Emis=E_No.get()
  BName=B_Name.get()
  Bbranch=Branch.get()
  Account=Account_No.get()
  Ifsc=IFSC.get()
  Micr=MICR.get()
  Umis=UMIS.get()
  NM=Naanmudhalvan.get()
  PP=Puthumaipen.get()
  TP=Tamilpudhalvan.get()

  if register=="" or name=="" or accyear=="" or Class=="" or gender=="" or dob=="" or Mobile=="" or Mail=="" or Blood=="" :
     Father=="" or Aadhaar=="" or FOccupation=="" or Mother=="" or MOccupation=="" or ParentNo=="" or Mtongue=="" or religion==""
     community=="" or SubCast=="" or income=="" or C_Address=="" or P_Address=="" or School=="" or S_Type=="" or medium==""
     Board=="" or mark=="" or percentage=="" or attempt=="" or Month=="" or year=="" or Emis=="" or BName=="" or Bbranch==""
     Account=="" or Ifsc=="" or Micr=="" or Umis=="" or NM=="" or PP=="" or TP==""

     Mail_regex=r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"
     if re.match(Mail_regex,Mail):
        messagebox.showinfo("Success",f"Email Address:{Mail} is valid.")
     else:
        messagebox.showerror("Error","Invalid Email Address ! please try again.")


     messagebox.showerror("error !","Few Data is Missing !")
      
  elif messagebox.showinfo("info","Successfully data Enterd !"):
       
     clear()
     
  
     file=load_workbook('Student_info.xlsx')
     sheet=file.active


  


  sheet.cell(column=1,row=sheet.max_row+1,value=register)
  sheet.cell(column=2,row=sheet.max_row,value=name)
  sheet.cell(column=3,row=sheet.max_row,value=accyear)
  sheet.cell(column=4,row=sheet.max_row,value=Class)
  sheet.cell(column=5,row=sheet.max_row,value=gender)
  sheet.cell(column=6,row=sheet.max_row,value=dob)
  sheet.cell(column=7,row=sheet.max_row,value=Mobile)
  sheet.cell(column=8,row=sheet.max_row,value=Mail)
  sheet.cell(column=9,row=sheet.max_row,value=Blood)
  sheet.cell(column=10,row=sheet.max_row,value=Aadhaar)
  sheet.cell(column=11,row=sheet.max_row,value=Father)
  sheet.cell(column=12,row=sheet.max_row,value=FOccupation)
  sheet.cell(column=13,row=sheet.max_row,value=Mother)
  sheet.cell(column=14,row=sheet.max_row,value=MOccupation)
  sheet.cell(column=15,row=sheet.max_row,value=ParentNo)
  sheet.cell(column=16,row=sheet.max_row,value=Mtongue)
  sheet.cell(column=17,row=sheet.max_row,value=religion)
  sheet.cell(column=18,row=sheet.max_row,value=community)
  sheet.cell(column=19,row=sheet.max_row,value=SubCast)
  sheet.cell(column=20,row=sheet.max_row,value=income)
  sheet.cell(column=21,row=sheet.max_row,value=C_Address)
  sheet.cell(column=22,row=sheet.max_row,value=P_Address)
  sheet.cell(column=23,row=sheet.max_row,value=School)
  sheet.cell(column=24,row=sheet.max_row,value=S_Type)
  sheet.cell(column=25,row=sheet.max_row,value=medium)
  sheet.cell(column=26,row=sheet.max_row,value=Board)
  sheet.cell(column=27,row=sheet.max_row,value=mark)
  sheet.cell(column=28,row=sheet.max_row,value=percentage)
  sheet.cell(column=29,row=sheet.max_row,value=attempt)
  sheet.cell(column=30,row=sheet.max_row,value=Month)
  sheet.cell(column=31,row=sheet.max_row,value=year)
  sheet.cell(column=32,row=sheet.max_row,value=Emis)
  sheet.cell(column=33,row=sheet.max_row,value=BName)
  sheet.cell(column=34,row=sheet.max_row,value=Bbranch)
  sheet.cell(column=35,row=sheet.max_row,value=Account)
  sheet.cell(column=36,row=sheet.max_row,value=Ifsc)
  sheet.cell(column=37,row=sheet.max_row,value=Micr)
  sheet.cell(column=38,row=sheet.max_row,value=Umis)
  sheet.cell(column=39,row=sheet.max_row,value=NM)
  sheet.cell(column=40,row=sheet.max_row,value=PP)
  sheet.cell(column=41,row=sheet.max_row,value=TP)


  
  
  
  
  

  file.save(r'Student_info.xlsx')
  




def clear():
   Register.set('')
   Name.set('')
   Accademicyear.set('')
   Class_Combobox.set('Select class')
   gender_Combobox.set('Select Gender')
   M_tongue.set('')
   Nationality.set('')
   F_Name.set('')
   M_Name.set('')
   Parent_No.set('')
   Community.set('')
   C_AddressEntry.delete(1.0,END)
   DOB.set('')
   Mobile_No.set('')
   Mail_Id.set('')
   B_group.set('')
   Religion.set('')
   F_Occupation.set('')
   M_Occupation.set('')
   Income.set('')
   Sub_Cast.set('')
   P_AddressEntry.delete(1.0,END)
   S_Name.set('')
   S_Type_Combobox.set('Select Type')
   Medium.set('')
   Board_Combobox.set('Select Board')
   Mark.set('')
   Percentage.set('')
   Attempt.set('')
   P_Month.set('')
   P_year.set('')
   E_No.set('')
   B_Name.set('')
   Branch.set('')
   Account_No.set('')
   IFSC.set('')
   MICR.set('')
   UMIS.set('')
   Naanmudhalvan.set('')
   Puthumaipen.set('')
   Tamilpudhalvan.set('')
   Aadhar.set('')
   


def search():
    reg_number = search_entry.get().strip()  
    
    if not reg_number:
        messagebox.showerror("Input Error", "Please enter a registration number.")
        return

    path = "C:\\Users\\ELCOT\\Desktop\\SRS(Studet Registration System)\\Student_info.xlsx"  
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    found = False
    student_details = ""

    
    for row in sheet_obj.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == reg_number:  
            student_details = (

                f"Register No""\t\t"f": {row[0]}"        "\t\t\t"f"Religion""\t\t\t"f": {row[16]}\n"  
                f"Name""\t\t\t"f": {row[1]}"             "\t\t\t"f"Community""\t\t"f": {row[17]}\n" 
                f"Academic Year""\t\t"f": {row[2]}"      "\t\t\t"f"Sub Cast""\t\t\t"f": {row[18]}\n"                  
                f"Class""\t\t\t"f": {row[3]}"            "\t\t\t\t"f"Income""\t\t\t"f": {row[19]}\n"         
                f"Gender""\t\t\t"f": {row[4]}\n"          
                f"DOB""\t\t\t"f": {row[5]}"              "\t\t\t"f"Bank Name""\t\t"f": {row[32]}\n"
                f"Mobile no""\t\t"f": {row[6]}"          "\t\t\t"f"Bank Branch""\t\t"f": {row[33]}\n"
                f"Mail id""\t\t\t"f": {row[7]}"          "\t\t"f"Account NO""\t\t"f": {row[34]}\n"
                f"Blood group""\t\t"f": {row[8]}"        "\t\t\t\t"f"IFSC Code""\t\t"f": {row[35]}\n"  
                f"Aadhar No""\t\t"f": {row[9]}"          "\t\t\t"f"MICR Code""\t\t"f": {row[36]}\n"
                f"Father Name""\t\t"f": {row[10]}\n"      
                f"Father's Occupation""\t"f": {row[11]}\n"
                f"Mother Name""\t\t"f": {row[12]}"      "\t\t\t"f"UMIS No""\t\t\t"f": {row[37]}\n"
                f"Mother's Occupation""\t"f": {row[13]}""\t\t\t"f"NM ID""\t\t\t"f": {row[38]}\n"
                f"Parent's Mobile No""\t"f": {row[14]}" "\t\t\t"f"PP ""\t\t\t"f": {row[39]}\n"
                f"Mother Tongue""\t\t"f": {row[15]}"    "\t\t\t\t"f"TP ""\t\t\t"f": {row[40]}\n"

                f"Communication Add""\t"f": {row[20]}\n"
                f"Permanent Add""\t\t"f": {row[21]}\n"
                f"School name""\t\t"f": {row[22]}\n"
                f"School Type""\t\t"f": {row[23]}\n"
                f"Medium""\t\t\t"f": {row[24]}\n"
                f"Board""\t\t\t"f": {row[25]}\n"
                f"12th Mark""\t\t"f": {row[26]}\n"
                f"Percentage""\t\t"f": {row[27]}\n"
                f"NO of Attempt""\t\t"f": {row[28]}\n"
                f"Passing month""\t\t"f": {row[29]}\n"
                f"Passing year""\t\t"f": {row[30]}\n"
                f"EMIS NO""\t\t\t"f": {row[31]}\n"
                
      
                
                
                
            )
            found = True
            break
        
        
        

    if found:
        messagebox.showinfo("Student Details", student_details)
    else:
        messagebox.showwarning("Not Found", "No student found with that registration number.")
        
search_entry = Entry(page2, width=15,bd=1,font="arial 20 ")
search_entry.place(x=950,y=78)

Button(page2, text="Search",width=10,height=1,bg="white",font="arial 13 ", command=search).place(x=1200,y=80)



#=========================================================================================
   

Label(page2,text="Register No:",font="arial 14",fg=framefg).place(x=400,y=150)
Label(page2,text="*",font="arial 13",fg="red").place(x=510,y=155)

Register=StringVar()
reg_entry =Entry(page2,textvariable=Register,width=17,font="arial 13")
reg_entry.place(x=535,y=153)

Label(page2,text="Accademic Year",font="arial 14",fg=framefg).place(x=1127,y=370)
Label(page2,text="Ex:( 2022 - 2025 )",font="arial 9",fg=framefg).place(x=1145,y=395)
Label(page2,text="*",font="arial 14",fg="red").place(x=1190,y=410)

Accademicyear=StringVar()
acc_entry =Entry(page2,textvariable=Accademicyear,width=17,font="arial 15")
acc_entry.place(x=1105,y=435)

Label(page2,text="Date:",width=5,height=1,font="arial 13 bold",fg='white',bg='#326273').place(x=250,y=85)
Date = StringVar()


today = date.today()
d1 = today.strftime("%d/%m/%y")
date_entry = Entry(page2,textvariable=Date,width=15,font="arial 10")
date_entry.place(x=300,y=86)
Date.set(d1)

#label

pg2_obj=LabelFrame(page2,text="Student's Details",font=20,bd=2,width=960,bg=framebg,fg=framefg,height=505,relief=GROOVE)
pg2_obj.place(x=60,y=203)

Label(pg2_obj,text="Full Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=20)
Label(pg2_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=165,y=25)
Label(pg2_obj,text="Class:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=65)
Label(pg2_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=165,y=60)
Label(pg2_obj,text="Gender:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)
Label(pg2_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=165,y=105)
Label(pg2_obj,text="Mother tongue:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=140)
Label(pg2_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=165,y=145)
Label(pg2_obj,text="Nationality:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=180)
Label(pg2_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=165,y=185)

Label(pg2_obj,text="Father Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=220)
Label(pg2_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=165,y=225)
Label(pg2_obj,text="Mother Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=260)
Label(pg2_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=165,y=265)
Label(pg2_obj,text="Parent Mobile NO",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=300)
Label(pg2_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=165,y=305)
Label(pg2_obj,text="Community:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=340)
Label(pg2_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=165,y=345)
Label(pg2_obj,text="Communication",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=380)
Label(pg2_obj,text="Address:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=400)
Label(pg2_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=165,y=407)

Label(pg2_obj,text="Date of Birth:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=20)
Label(pg2_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=655,y=25)
Label(pg2_obj,text="Mobile No:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=60)
Label(pg2_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=655,y=65)
Label(pg2_obj,text="Email ID:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)
Label(pg2_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=655,y=105)
Label(pg2_obj,text="Blood Group:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=140)
Label(pg2_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=655,y=145)
Label(pg2_obj,text="Religion:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=180)
Label(pg2_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=655,y=185)

Label(pg2_obj,text="Father's Occupation:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=220)
Label(pg2_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=655,y=225)
Label(pg2_obj,text="Mother's Occupation:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=260)
Label(pg2_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=655,y=265)
Label(pg2_obj,text="Income:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=300)
Label(pg2_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=655,y=305)
Label(pg2_obj,text="Sub Cast:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=340)
Label(pg2_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=655,y=345)
Label(pg2_obj,text="Permanent Address:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=380)
Label(pg2_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=655,y=385)





Name=StringVar()
name_entry = Entry(pg2_obj,textvariable=Name,width=30,font="arial 10")
name_entry.place(x=185,y=20)

Class_Combobox= Combobox(pg2_obj,values=['I BCA','II BCA','III BCA'],font="Roboto 10",width=27,state="r")
Class_Combobox.place(x=185,y=60)
Class_Combobox.set("Select class")

#gender

gender_Combobox= Combobox(pg2_obj,values=['Male','Female'],font="Roboto 10",width=27,state="r")
gender_Combobox.place(x=185,y=100)
gender_Combobox.set("Select Gender")

#==========


M_tongue=StringVar()
mtongue_entry = Entry(pg2_obj,textvariable=M_tongue,width=30,font="arial 10")
mtongue_entry.place(x=185,y=140)

Nationality=StringVar()
nationality_entry = Entry(pg2_obj,textvariable=Nationality,width=30,font="arial 10")
nationality_entry.place(x=185,y=180)

#===============================

F_Name=StringVar()
f_name_entry = Entry(pg2_obj,textvariable=F_Name,width=30,font="arial 10")
f_name_entry.place(x=185,y=220)

M_Name=StringVar()
m_name_entry = Entry(pg2_obj,textvariable=M_Name,width=30,font="arial 10")
m_name_entry.place(x=185,y=260)
#=================================

def validate_num(action,value):
    if action =="1":
        return value.isdigit()and len(value)<=10
    return True
validate_cmd = pg2_obj.register(validate_num)
Parent_No=StringVar()
pno_entry = Entry(pg2_obj,textvariable=Parent_No,validate="key",width=30,font="arial 10",validatecommand=(validate_cmd,"%d","%P"))
pno_entry.place(x=185,y=300)
#=================================

Community=StringVar()
community_entry = Entry(pg2_obj,textvariable=Community,width=30,font="arial 10")
community_entry.place(x=185,y=340)

C_AddressEntry = Text(page2,width=30,height=4,bd=1,font="arial 10")
C_AddressEntry.place(x=246,y=600)

#========================

def submit_dob():
    dob=dob_entry.get()
    

DOB=StringVar()
dob_entry = DateEntry(pg2_obj,textvariable=DOB,width=24,bg="darkblue",fg="white",date_pattern="dd-mm-yyyy")
dob_entry.place(x=675,y=20)

#====================
def validate_num(action,value):
    if action =="1":
        return value.isdigit()and len(value)<=10
    return True
validate_cmd = pg2_obj.register(validate_num)
Mobile_No=StringVar()
mobileno_entry = Entry(pg2_obj,textvariable=Mobile_No,validate="key",width=30,font="arial 10",validatecommand=(validate_cmd,"%d","%P"))
mobileno_entry.place(x=675,y=60)
#============================
def validate_email(action,value):
    if action =="1":
        return True
    return True


validate_cmd = pg2_obj.register(validate_email)                            
Mail_Id=StringVar()
mailid_entry = Entry(pg2_obj,textvariable=Mail_Id,validate="key",width=30,font="arial 10",validatecommand=(validate_cmd,"%d","%P"))
mailid_entry.place(x=675,y=100)

#================================

B_group=StringVar()
bgroup_entry = Entry(pg2_obj,textvariable=B_group,width=30,font="arial 10")
bgroup_entry.place(x=675,y=140)

Religion=StringVar()
religion_entry = Entry(pg2_obj,textvariable=Religion,width=30,font="arial 10")
religion_entry.place(x=675,y=180)


#==================================

F_Occupation=StringVar()
f_occupation_entry = Entry(pg2_obj,textvariable=F_Occupation,width=30,font="arial 10")
f_occupation_entry.place(x=675,y=220)

M_Occupation=StringVar()
m_occupation_entry = Entry(pg2_obj,textvariable=M_Occupation,width=30,font="arial 10")
m_occupation_entry.place(x=675,y=260)

Income=StringVar()
income_entry = Entry(pg2_obj,textvariable=Income,width=30,font="arial 10")
income_entry.place(x=675,y=300)

Sub_Cast=StringVar()
subcast_entry = Entry(pg2_obj,textvariable=Sub_Cast,width=30,font="arial 10")
subcast_entry.place(x=675,y=340)

P_AddressEntry = Text(page2,width=30,height=4,bd=1,font="arial 10")
P_AddressEntry.place(x=737,y=600)

#==================
#====IMAGE FRAME ==

f=Frame(page2,bd=2,bg="white",width=200,height=200,relief=GROOVE)
f.place(x=1100,y=150)

img2=PhotoImage(file="images/BCA logo.png")
lbl2=Label(f,bg="white",image=img2)
lbl2.place(x=-2,y=-2)

#===========================
#=========BUTTONS===========
page2_button= Button(page2,text="HOME",width=10,height=1,font="arial 12 bold",bg="light blue",bd=1, command=lambda:show_frame(page1))
page2_button.place(x=70,y=80)

page2_button= Button(page2,text="NEXT >",width=19,height=2,font="arial 12 bold",bg="Aquamarine",bd=1, command=lambda: show_frame(page3))
page2_button.place(x=1100,y=480)


page2_button= Button(page2,text="CLEAR",width=19,height=2,font="arial 12 bold",bg="yellow",bd=1,command=clear)
page2_button.place(x=1100,y=540)
  
     
def Exit():
    root.destroy()

Button(page2,text="EXIT",width=19,height=2,font="arial 12 bold",bg="Deep Pink",bd=1,command=Exit).place(x=1100,y=670)





#=========PAGE 3================


page3_label =Label(page3,text="Email : kamalcrush15@gmail.com ",width=10,height=3,bg="#778899",anchor='e').pack(side=TOP,fill=X)
page3_label =Label(page3,text="STUDENT REGISTRATION ",width=10,height=2,bg='#326273',fg='#fff',font='arial 20 bold').pack(side=TOP,fill=X)





Label(page3,text="Date:",width=5,height=1,font="arial 13 bold",fg='white',bg='#326273').place(x=250,y=85)
Date = StringVar()


today = date.today()
d1 = today.strftime("%d/%m/%y")
date_entry = Entry(page3,textvariable=Date,width=15,font="arial 10")
date_entry.place(x=300,y=86)
Date.set(d1)

pg3_obj=LabelFrame(page3,text="Student's Details",font=20,bd=2,width=960,bg=framebg,fg=framefg,height=505,relief=GROOVE)
pg3_obj.place(x=60,y=203)




Label(page3,text="NOTE: Only eligible students should fill ' YES ' for Puthumaipen & Tamilpudhalvan ,otherwise enter ' NO ' "
      ,font="arial 10 bold",fg=framefg).place(x=70,y=150)

Label(pg3_obj,text="School Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=20)
Label(pg3_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=150,y=20)
Label(pg3_obj,text="School Type:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=60)
Label(pg3_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=150,y=60)
Label(pg3_obj,text="Medium:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)
Label(pg3_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=150,y=100)
Label(pg3_obj,text="Board :",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=140)
Label(pg3_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=150,y=140)
Label(pg3_obj,text="12th Mark:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=180)
Label(pg3_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=150,y=180)

Label(pg3_obj,text="Percentage:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=220)
Label(pg3_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=150,y=220)
Label(pg3_obj,text="No of Attempt:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=260)
Label(pg3_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=150,y=260)
Label(pg3_obj,text="Passing Month:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=300)
Label(pg3_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=150,y=300)
Label(pg3_obj,text="Passing Year:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=340)
Label(pg3_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=150,y=340)
Label(pg3_obj,text="EMIS NO:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=380)
Label(pg3_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=150,y=380)

Label(pg3_obj,text="Bank Name:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=20)
Label(pg3_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=645,y=25)
Label(pg3_obj,text="Bank Branch:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=60)
Label(pg3_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=645,y=65)
Label(pg3_obj,text="Account No:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)
Label(pg3_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=645,y=105)
Label(pg3_obj,text="IFSC Code:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=140)
Label(pg3_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=645,y=145)
Label(pg3_obj,text="MICR Code:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=180)
Label(pg3_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=645,y=185)

Label(pg3_obj,text="UMIS No:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=220)
Label(pg3_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=645,y=225)
Label(pg3_obj,text="Naanmudhalvan ID:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=260)
Label(pg3_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=645,y=265)
Label(pg3_obj,text="Puthumaipen:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=300)
Label(pg3_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=645,y=305)
Label(pg3_obj,text="Tamilpudhalvan:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=340)
Label(pg3_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=645,y=345)
Label(pg3_obj,text="Aadhar No:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=380)
Label(pg3_obj,text="*",font="arial 13",bg=framebg,fg="red").place(x=645,y=385)


S_Name=StringVar()
sname_entry = Entry(pg3_obj,textvariable=S_Name,width=30,font="arial 10")
sname_entry.place(x=175,y=20)

S_Type_Combobox= Combobox(pg3_obj,values=['GOVERNMENT','PRIVATE','GOVT AIDED'],font="Roboto 10",width=27,state="r")
S_Type_Combobox.place(x=175,y=60)
S_Type_Combobox.set("Select Type" )

Medium=StringVar()
medium_entry = Entry(pg3_obj,textvariable=Medium,width=30,font="arial 10")
medium_entry.place(x=175,y=100)



Board_Combobox= Combobox(pg3_obj,values=['State Board','CBSE'],font="Roboto 10",width=27,state="r")
Board_Combobox.place(x=175,y=140)
Board_Combobox.set("Select Board")


Mark=StringVar()
mark_entry = Entry(pg3_obj,textvariable=Mark,width=30,font="arial 10")
mark_entry.place(x=175,y=180)

Percentage=StringVar()
percentage_entry = Entry(pg3_obj,textvariable=Percentage,width=30,font="arial 10")
percentage_entry.place(x=175,y=220)

Attempt=StringVar()
attempt_entry = Entry(pg3_obj,textvariable=Attempt,width=30,font="arial 10")
attempt_entry.place(x=175,y=260)

P_Month=StringVar()
pmonth = Entry(pg3_obj,textvariable=P_Month,width=30,font="arial 10")
pmonth.place(x=175,y=300)

P_year=StringVar()
pyear_entry = Entry(pg3_obj,textvariable=P_year,width=30,font="arial 10")
pyear_entry.place(x=175,y=340)

E_No=StringVar()
eno_entry = Entry(pg3_obj,textvariable=E_No,width=30,font="arial 10")
eno_entry.place(x=175,y=380)



#=================================


B_Name=StringVar()
bname_entry = Entry(pg3_obj,textvariable=B_Name,width=30,font="arial 10")
bname_entry.place(x=665,y=20)


Branch=StringVar()
branch_entry = Entry(pg3_obj,textvariable=Branch,width=30,font="arial 10")
branch_entry.place(x=665,y=60)

Account_No=StringVar()
accountno_entry = Entry(pg3_obj,textvariable=Account_No,width=30,font="arial 10")
accountno_entry.place(x=665,y=100)

IFSC=StringVar()
ifsc_entry = Entry(pg3_obj,textvariable=IFSC,width=30,font="arial 10")
ifsc_entry.place(x=665,y=140)

MICR=StringVar()
micr_entry = Entry(pg3_obj,textvariable=MICR,width=30,font="arial 10")
micr_entry.place(x=665,y=180)

#==============================

UMIS=StringVar()
umis_entry = Entry(pg3_obj,textvariable=UMIS,width=30,font="arial 10")
umis_entry.place(x=665,y=220)

Naanmudhalvan=StringVar()
naanmudhalvan_entry = Entry(pg3_obj,textvariable=Naanmudhalvan,width=30,font="arial 10")
naanmudhalvan_entry.place(x=665,y=260)

Puthumaipen=StringVar()
puthumaipen_entry = Entry(pg3_obj,textvariable=Puthumaipen,width=30,font="arial 10")
puthumaipen_entry.place(x=665,y=300)

Tamilpudhalvan=StringVar()
tamilpudhalvan_entry = Entry(pg3_obj,textvariable=Tamilpudhalvan,width=30,font="arial 10")
tamilpudhalvan_entry.place(x=665,y=340)
#=========================
def validate_num(action,value):
    if action =="1":
        return value.isdigit()and len(value)<=12
    return True
validate_cmd = pg3_obj.register(validate_num)
Aadhar=StringVar()
aadhar_entry = Entry(pg3_obj,textvariable=Aadhar,validate="key",width=30,font="arial 10",validatecommand=(validate_cmd,"%d","%P"))
aadhar_entry.place(x=665,y=380)

#==================


saveButton=Button(page3,text="SUBMIT",width=19,height=2,font="arial 12 bold",bg="light green",bd=1,command=submit)
saveButton.place(x=1100,y=370)



f=Frame(page3,bd=2,bg="white",width=200,height=200,relief=GROOVE)
f.place(x=1100,y=150)

img3=PhotoImage(file="images/BCA logo.png")
lbl3=Label(f,bg="white",image=img3)
lbl3.place(x=-2,y=-2)

page3_button= Button(page3,text="HOME",width=10,height=1,font="arial 12 bold",bg="light blue",bd=1, command=lambda:show_frame(page1))
page3_button.place(x=70,y=80)

page3_button= Button(page3,text="< BACK ",width=19,height=2,font="arial 12 bold",bg="Aquamarine",bd=1, command=lambda: show_frame(page2))
page3_button.place(x=1100,y=480)

page3_button= Button(page3,text="CLEAR",width=19,height=2,font="arial 12 bold",bg="yellow",bd=1,command=clear)
page3_button.place(x=1100,y=540)





def Exit():
    root.destroy()

Button(page3,text="EXIT",width=19,height=2,font="arial 12 bold",bg="Deep Pink",bd=1,command=Exit).place(x=1100,y=670)




root.mainloop()
